# -*- coding: utf-8 -*-
"""
Скрипт для создания единого Excel файла проекта RedPeak.
Источники: Генеральный_план_работы_Redpeak_v4.xlsx + RedPeak_Roadmap_2026_2028.docx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os

# === КОНФИГУРАЦИЯ ===
OUTPUT_PATH = r"D:\REDPEAK\Agent systems\AgentHQ\DATA\Дорожная карта проекта REDPEAK\RedPeak_Единый_план_2026_2028.xlsx"
SOURCE_XLSX = r"C:\Users\sashatrash\OneDrive\Desktop\Генеральный_план_работы_Redpeak_v4.xlsx"

# === СТИЛИ ===
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
PHASE1_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # зелёный
PHASE2_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # синий
PHASE3_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # жёлтый
PHASE4_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # оранжевый
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# === ДАННЫЕ ===

def format_date(val):
    """Форматирование даты."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, date):
        return val.strftime("%d.%m.%Y")
    s = str(val)
    if "00:00:00" in s:
        try:
            dt = datetime.strptime(s.split(" ")[0], "%Y-%m-%d")
            return dt.strftime("%d.%m.%Y")
        except:
            pass
    return s


def get_phase_fill(phase):
    fills = {
        "Фаза 1": PHASE1_FILL,
        "Фаза 2": PHASE2_FILL,
        "Фаза 3": PHASE3_FILL,
        "Фаза 4": PHASE4_FILL,
    }
    return fills.get(phase, None)


def apply_row_style(ws, row_num, num_cols, font=NORMAL_FONT, fill=None, alignment=LEFT_WRAP):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = alignment
        if fill:
            cell.fill = fill


def apply_header(ws, row_num, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def auto_width(ws, min_width=8, max_width=50):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, len(line) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


# ================================================================
# ЛИСТ 1: ПЛАН РАЗВИТИЯ
# ================================================================
def create_plan_sheet(wb):
    ws = wb.active
    ws.title = "План развития"

    # Заголовок
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="ПЛАН РАЗВИТИЯ ПРОЕКТА REDPEAK 2026-2028")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    info_cell = ws.cell(row=2, column=1,
        value="Дата: 04.04.2026  |  Горизонт: Q1 2026 - Q4 2028  |  Собственные: 2 043 400 руб.  |  Грант ФСИ: 5 000 000 руб.")
    info_cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
    info_cell.alignment = Alignment(horizontal="center")

    # Шапка таблицы
    headers = ["No", "Фаза", "Задача", "Владелец", "Результат", "Срок", "Статус", "% гот."]
    apply_header(ws, 4, headers)

    # --- ФАЗА 1: ФУНДАМЕНТ (Q1-Q2 2026) ---
    row = 5
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="ФАЗА 1: ФУНДАМЕНТ (Q1-Q2 2026)")
    apply_row_style(ws, row, 8, font=SECTION_FONT, fill=PHASE1_FILL, alignment=CENTER)
    row += 1

    # Задачи из старого плана (Блок A + частично B, C, D)
    phase1_tasks = [
        ["1.2", "Фаза 1", "Прогон 50 деталей через базовый алгоритм + оценка качества", "Владислав",
         "Таблица: деталь - главный вид / доп. виды / сечение / оценка %", "14.02.2026", "В работе", "10%"],
        ["1.3", "Фаза 1", "Алгоритм очистки от подписей, штампов, децимальных номеров", "Владислав",
         "Утилита очистки v2 + отчёт о юридической чистоте", "17.02.2026", "В работе", "80%"],
        ["1.4", "Фаза 1", "Заявка на грант ФСИ «Старт-ЦТ»", "Владислав",
         "Заявка подана (PDF + приложения)", "01.03.2026", "В работе", "80%"],
        ["1.5", "Фаза 1", "Подготовить топ 200 пар деталь/чертёж для пилота", "Владислав",
         "Подготовленные пары", "10.02.2026", "Завершено", "100%"],
        ["1.6", "Фаза 1", "Найти нового ML-инженера для пилота AI-агентов", "Владислав",
         "Кандидат найден + договор", "30.04.2026", "Завершено", "100%"],
        ["2.3", "Фаза 1", "CAD-разработчик: экстракция признаков (S1-S3)", "Алексей (CAD)",
         "features.csv + rules_decision.json + разметка.csv", "30.04.2026", "Ожидание договора", "0%"],
        ["2.4", "Фаза 1", "Прогон 500 деталей (расширенное тестирование)", "Владислав + Алексей",
         "Полная таблица coverage на 500 деталях", "31.03.2026", "Не начато", "0%"],
        ["3.1", "Фаза 1", "Выход в найм (Solidworks)", "Владислав",
         "Работа найдена; фокус на проект сохранён", "07.03.2026", "Завершено", "100%"],
        ["4.1", "Фаза 1", "Расширить датасет с 13 000 до 20 000 пар", "Владислав",
         "Датасет 20 000 пар (m3d + cdw)", "30.04.2026", "Не начато", "0%"],
        # Новые из roadmap Фаза 1
        ["R1.1", "Фаза 1", "IT-аккредитация Минцифры", "РП",
         "Налоговые льготы (5% прибыль, 15% взносы)", "Апрель 2026", "Не начато", "0%"],
        ["R1.2", "Фаза 1", "Резидентство Сколково (формальная проверка)", "РП",
         "Доступ к грантам + 0% налог на прибыль", "Q2 2026", "В работе", ""],
        ["R1.3", "Фаза 1", "Облачные гранты (VK Cloud, Yandex)", "РП",
         "До 3 млн руб. на GPU-ресурсы", "Май 2026", "Не начато", "0%"],
        ["R1.4", "Фаза 1", "Резидентство инкубатора «Ингрия»", "РП",
         "Менторство + связи с инвесторами", "Июнь 2026", "Не начато", "0%"],
    ]

    for task in phase1_tasks:
        for col, val in enumerate(task, 1):
            ws.cell(row=row, column=col, value=val)
        apply_row_style(ws, row, 8, fill=PHASE1_FILL if "R" in str(task[0]) else None)
        row += 1

    # --- ФАЗА 2: MVP (Q3-Q4 2026) ---
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="ФАЗА 2: MVP И ПЕРВЫЕ ПОЛЬЗОВАТЕЛИ (Q3-Q4 2026)")
    apply_row_style(ws, row, 8, font=SECTION_FONT, fill=PHASE2_FILL, alignment=CENTER)
    row += 1

    phase2_tasks = [
        ["5.1", "Фаза 2", "MVP-0: главный вид + доп. виды + сечение", "Алексей (CAD)",
         "CDW: главный вид + доп. виды + сечение + ГОСТ", "15.05.2026", "Не начато", "0%"],
        ["5.2", "Фаза 2", "MVP-1: правила + ML (адаптивно по coverage)", "Алексей + ML-инженер",
         "CDW с оптимальными видами + интеллектуальные разрезы", "30.06.2026", "Не начато", "0%"],
        ["5.3", "Фаза 2", "Расширенный функционал: выносные элементы + размеры", "Алексей + ML-инженер",
         "CDW с выносными элементами, размерами, оформление по ГОСТ", "31.10.2026", "Не начато", "0%"],
        ["5.4", "Фаза 2", "Интеграционное тестирование + стабилизация RC", "Команда",
         "0 критических багов; бесперебойная работа на 100 деталях", "15.07.2026", "Не начато", "0%"],
        ["4.2", "Фаза 2", "Тарифная сетка + landing page", "Владислав",
         "Тарифы валидированы с 3+ клиентами", "31.05.2026", "Не начато", "0%"],
        # Новые из roadmap Фаза 2
        ["R2.1", "Фаза 2", "Pipeline F1-F7: стабильный прогон на 50 деталях", "Разработка",
         "Coverage >70% - MVP на правилах", "Август 2026", "Не начато", "0%"],
        ["R2.2", "Фаза 2", "Точка решения: правила vs ML (неделя 6)", "Команда",
         "Выбор архитектурного сценария A/B/C", "Август 2026", "Не начато", "0%"],
        ["R2.3", "Фаза 2", "Микрогрант Сколково на пилоты", "РП",
         "До 1,5 млн руб. на внедрение", "Октябрь 2026", "Не начато", "0%"],
        ["R2.4", "Фаза 2", "Регистрация ПО в Роспатенте", "РП",
         "Защита ИС (требование гранта ФСИ)", "Ноябрь 2026", "Не начато", "0%"],
        ["R2.5", "Фаза 2", "3-5 пилотных внедрений", "РП + Продажи",
         "Реальные кейсы и обратная связь", "Ноябрь 2026", "Не начато", "0%"],
        ["R2.6", "Фаза 2", "MVP-1: полный продукт + документация", "Разработка",
         "Релиз для тестовых продаж", "Декабрь 2026", "Не начато", "0%"],
    ]

    for task in phase2_tasks:
        for col, val in enumerate(task, 1):
            ws.cell(row=row, column=col, value=val)
        apply_row_style(ws, row, 8, fill=PHASE2_FILL if "R" in str(task[0]) else None)
        row += 1

    # --- ФАЗА 3: РОСТ (2027) ---
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="ФАЗА 3: ТЕСТОВЫЕ ПРОДАЖИ И РОСТ (2027)")
    apply_row_style(ws, row, 8, font=SECTION_FONT, fill=PHASE3_FILL, alignment=CENTER)
    row += 1

    phase3_tasks = [
        ["R3.1", "Фаза 3", "Включение в реестр отечественного ПО", "РП",
         "Приоритет в госзакупках + 0% НДС", "Q1 2027", "Не начато", "0%"],
        ["R3.2", "Фаза 3", "Запуск продаж через каталог АСКОН", "Продажи",
         "Доступ к 13 000 предприятий", "Q1 2027", "Не начато", "0%"],
        ["R3.3", "Фаза 3", "Листинг на Softline Store", "Продажи",
         "B2B-маркетплейс, 2000+ менеджеров", "Q1 2027", "Не начато", "0%"],
        ["R3.4", "Фаза 3", "Заявка на ФСИ «Старт-2» (10 млн руб.)", "РП",
         "Финансирование второго этапа R&D", "Q2 2027", "Не начато", "0%"],
        ["R3.5", "Фаза 3", "Интеграция с nanoCAD / T-FLEX", "Разработка",
         "Мультиплатформенность - x2 рынок", "Q3 2027", "Не начато", "0%"],
        ["R3.6", "Фаза 3", "Конференции: ЦИПР, KOMPAScon, ИИПром", "РП",
         "Узнаваемость, лиды, партнёры", "2027", "Не начато", "0%"],
        ["R3.7", "Фаза 3", "1 000 проданных лицензий", "Продажи",
         "Выручка ~8,4 млн руб.", "Q4 2027", "Не начато", "0%"],
        ["R3.8", "Фаза 3", "ФРИИ / венчурный раунд (опционально)", "РП + НР",
         "До 25 млн руб. инвестиций", "Q3-Q4 2027", "Не начато", "0%"],
    ]

    for task in phase3_tasks:
        for col, val in enumerate(task, 1):
            ws.cell(row=row, column=col, value=val)
        apply_row_style(ws, row, 8, fill=PHASE3_FILL)
        row += 1

    # --- ФАЗА 4: МАСШТАБ (2028) ---
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="ФАЗА 4: МАСШТАБИРОВАНИЕ (2028)")
    apply_row_style(ws, row, 8, font=SECTION_FONT, fill=PHASE4_FILL, alignment=CENTER)
    row += 1

    phase4_tasks = [
        ["R4.1", "Фаза 4", "Обновление продукта: разрезы + авторасстановка размеров", "Разработка",
         "Этапы 3-4 продукта", "Q1 2028", "Не начато", "0%"],
        ["R4.2", "Фаза 4", "ФСИ «Бизнес-Старт» (12 млн руб.)", "РП",
         "Финансирование масштабирования", "Q1 2028", "Не начато", "0%"],
        ["R4.3", "Фаза 4", "Участие в госзакупках (44-ФЗ, 223-ФЗ)", "Продажи",
         "Ростех, ОАК, ОСК, РЖД", "Q1-Q2 2028", "Не начато", "0%"],
        ["R4.4", "Фаза 4", "Экспансия: Беларусь, Казахстан", "Продажи",
         "+30% адресного рынка (СНГ)", "Q2-Q3 2028", "Не начато", "0%"],
        ["R4.5", "Фаза 4", "7 500 проданных лицензий", "Продажи",
         "Выручка ~66,8 млн руб.", "Q4 2028", "Не начато", "0%"],
        ["R4.6", "Фаза 4", "Обновление продукта v2.0", "Разработка",
         "Полный цикл: модель - чертёж - спецификация", "Q4 2028", "Не начато", "0%"],
    ]

    for task in phase4_tasks:
        for col, val in enumerate(task, 1):
            ws.cell(row=row, column=col, value=val)
        apply_row_style(ws, row, 8, fill=PHASE4_FILL)
        row += 1

    # Настройка ширины
    col_widths = {"A": 7, "B": 10, "C": 50, "D": 20, "E": 45, "F": 16, "G": 16, "H": 8}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Закрепить шапку
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "2F5496"

    return ws


# ================================================================
# ЛИСТ 2: ЖУРНАЛ РАСХОДОВ
# ================================================================
def create_expenses_sheet(wb):
    ws = wb.create_sheet("Журнал расходов")

    # Читаем данные из исходного файла
    src_wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    src_ws = src_wb["Журнал расходов"]

    # Заголовок
    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1, value="ЖУРНАЛ РАСХОДОВ ПРОЕКТА REDPEAK")
    title.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title.alignment = Alignment(horizontal="center")

    # Шапка
    headers = ["Дата", "Категория", "Описание", "Сумма (руб.)", "Источник", "Связ. задача", "Примечание"]
    apply_header(ws, 3, headers)

    # Данные из исходника (строки 4-27)
    row = 4
    total = 0
    for src_row in range(4, src_ws.max_row + 1):
        date_val = src_ws.cell(row=src_row, column=1).value
        category = src_ws.cell(row=src_row, column=2).value
        description = src_ws.cell(row=src_row, column=3).value
        amount = src_ws.cell(row=src_row, column=4).value
        source = src_ws.cell(row=src_row, column=5).value
        task_link = src_ws.cell(row=src_row, column=6).value
        note = src_ws.cell(row=src_row, column=9).value

        # Пропустить пустые строки
        if not description and not amount:
            continue

        ws.cell(row=row, column=1, value=format_date(date_val) if date_val else "")
        ws.cell(row=row, column=2, value=str(category) if category else "")
        ws.cell(row=row, column=3, value=str(description) if description else "")

        if amount:
            try:
                amt = float(amount)
                ws.cell(row=row, column=4, value=amt)
                ws.cell(row=row, column=4).number_format = '#,##0'
                total += amt
            except (ValueError, TypeError):
                ws.cell(row=row, column=4, value=amount)

        ws.cell(row=row, column=5, value=str(source) if source else "")
        ws.cell(row=row, column=6, value=str(task_link) if task_link else "")
        ws.cell(row=row, column=7, value=str(note) if note else "")

        apply_row_style(ws, row, 7)
        row += 1

    # Строка ИТОГО
    row += 1
    ws.cell(row=row, column=1, value="ИТОГО:")
    ws.cell(row=row, column=4, value=total)
    ws.cell(row=row, column=4).number_format = '#,##0'
    apply_row_style(ws, row, 7, font=TOTAL_FONT, fill=TOTAL_FILL)

    src_wb.close()

    # Ширина колонок
    col_widths = {"A": 14, "B": 28, "C": 50, "D": 14, "E": 14, "F": 12, "G": 30}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = "548235"

    return ws


# ================================================================
# ЛИСТ 3: БЮДЖЕТ
# ================================================================
def create_budget_sheet(wb):
    ws = wb.create_sheet("Бюджет")

    # Заголовок
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="БЮДЖЕТ ПРОЕКТА REDPEAK")
    title.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title.alignment = Alignment(horizontal="center")

    # Шапка
    headers = ["Статья расходов", "Бюджет (руб.)", "Факт (руб.)", "Остаток (руб.)", "Источник", "Примечание"]
    apply_header(ws, 3, headers)

    # === СОБСТВЕННЫЕ СРЕДСТВА ===
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="СОБСТВЕННЫЕ СРЕДСТВА")
    apply_row_style(ws, row, 6, font=SECTION_FONT, fill=SECTION_FILL, alignment=CENTER)
    row += 1

    own_items = [
        ["CAD-разработчик (базовый)", 300000, None, None, "Собственные", "Алексей К. + новый разработчик"],
        ["ML-инженер (пилот)", 200000, None, None, "Собственные", "Поиск нового ML-инженера"],
        ["Инфраструктура/серверы", 200000, None, None, "Собственные", "Серверы, офис, лицензии, юр. адрес, ЭДО"],
        ["Маркетинг/коммерциализация", 170000, None, None, "Собственные", "Упаковка, бренд, презентации"],
        ["Резерв/прочее", 1253400, None, None, "Собственные", "Консультанты, заявки Сколково/ФСИ, прочее"],
        ["Налоги и взносы", 97200, None, None, "Собственные", "Страховые взносы 30% от МРОТ (с 01.01.2026)"],
    ]

    own_total_budget = 0
    for item in own_items:
        for col, val in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in (2, 3, 4) and val is not None:
                cell.number_format = '#,##0'
        if item[1]:
            own_total_budget += item[1]
        # Формула остатка = бюджет - факт
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = '#,##0'
        apply_row_style(ws, row, 6)
        row += 1

    # Итого собственные
    ws.cell(row=row, column=1, value="Итого собственные средства")
    ws.cell(row=row, column=2, value=f"=SUM(B5:B{row-1})")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3, value=f"=SUM(C5:C{row-1})")
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    ws.cell(row=row, column=4).number_format = '#,##0'
    apply_row_style(ws, row, 6, font=TOTAL_FONT, fill=TOTAL_FILL)
    own_total_row = row
    row += 2

    # === ГРАНТ ФСИ ===
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="ГРАНТ ФСИ «СТАРТ-ЦТ» (5 000 000 руб.) - ещё не получен")
    apply_row_style(ws, row, 6, font=SECTION_FONT, fill=SECTION_FILL, alignment=CENTER)
    row += 1
    grant_start_row = row

    grant_items = [
        ["CAD-разработчик (грант)", 1900000, None, None, "Грант ФСИ", "Phase 2: выносные, размеры"],
        ["ML-инженер (грант)", 1850000, None, None, "Грант ФСИ", "Обучение ML, инференс"],
        ["Инфраструктура/серверы (грант)", 150000, None, None, "Грант ФСИ", "Вычислительные ресурсы ML"],
        ["Тестирование/QA", 400000, None, None, "Грант ФСИ", "Нагрузочные тесты, верификация"],
        ["Маркетинг (грант)", 300000, None, None, "Грант ФСИ", "Лендинг, демо, продажи"],
        ["Резерв (грант)", 400000, None, None, "Грант ФСИ", "8% резерв"],
    ]

    for item in grant_items:
        for col, val in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in (2, 3, 4) and val is not None:
                cell.number_format = '#,##0'
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=4).number_format = '#,##0'
        apply_row_style(ws, row, 6)
        row += 1

    # Итого грант
    ws.cell(row=row, column=1, value="Итого грант ФСИ")
    ws.cell(row=row, column=2, value=f"=SUM(B{grant_start_row}:B{row-1})")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3, value=f"=SUM(C{grant_start_row}:C{row-1})")
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    ws.cell(row=row, column=4).number_format = '#,##0'
    apply_row_style(ws, row, 6, font=TOTAL_FONT, fill=TOTAL_FILL)
    grant_total_row = row
    row += 2

    # === ОБЩИЙ ИТОГ ===
    ws.cell(row=row, column=1, value="ОБЩИЙ БЮДЖЕТ (собственные + грант)")
    ws.cell(row=row, column=2, value=f"=B{own_total_row}+B{grant_total_row}")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3, value=f"=C{own_total_row}+C{grant_total_row}")
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    ws.cell(row=row, column=4).number_format = '#,##0'
    apply_row_style(ws, row, 6, font=Font(name="Calibri", bold=True, size=12), fill=PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"))

    # Ширина колонок
    col_widths = {"A": 35, "B": 16, "C": 16, "D": 16, "E": 14, "F": 40}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = "C00000"

    return ws


# ================================================================
# ЛИСТ 4: ОПЕРАТИВНЫЕ ЗАДАЧИ
# ================================================================
def create_operative_sheet(wb):
    ws = wb.create_sheet("Оперативные задачи", 0)  # первый лист

    RED_HEADER = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    URGENT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    NORMAL_FILL_OP = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    LOW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Заголовок
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="ОПЕРАТИВНЫЕ ЗАДАЧИ")
    title.font = Font(name="Calibri", bold=True, size=14, color="C00000")
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    info = ws.cell(row=2, column=1, value="Актуально на: 04.04.2026")
    info.font = Font(name="Calibri", size=10, italic=True, color="595959")
    info.alignment = Alignment(horizontal="center")

    # Шапка
    headers = ["No", "Задача", "Детали / подзадачи", "Приоритет", "Дедлайн", "Статус"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = RED_HEADER
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    tasks = [
        {
            "no": "1",
            "task": "Поиск работы / трудоустройство",
            "details": "Отклики на hh.ru (автоматические + ручные по целевым компаниям). Хождение по собеседованиям.",
            "priority": "Высокий",
            "deadline": "Постоянно",
            "status": "В работе",
            "fill": URGENT_FILL,
        },
        {
            "no": "2",
            "task": "Проверить данные Алексея по этапу 2 ТЗ",
            "details": "Алексей сдал результаты по второму этапу разработки. Провести приёмку/проверку.",
            "priority": "Высокий",
            "deadline": "05.04.2026",
            "status": "Не начато",
            "fill": URGENT_FILL,
        },
        {
            "no": "3",
            "task": "Отправить в ФНС сведения об изменении юр. адреса",
            "details": "Заполнить документы + отправить через ЭП в налоговую.",
            "priority": "Высокий",
            "deadline": "05.04.2026",
            "status": "Не начато",
            "fill": URGENT_FILL,
        },
        {
            "no": "4",
            "task": "Оценить тех. документы по реверс-инжинирингу от потенциального работодателя",
            "details": "Изучить ТЗ/документацию. Оценить возможность удалённой работы и свои компетенции.",
            "priority": "Средний",
            "deadline": "На этой неделе",
            "status": "Не начато",
            "fill": NORMAL_FILL_OP,
        },
        {
            "no": "5",
            "task": "Практика: SolidWorks + КОМПАС-3D",
            "details": "Вспомнить SolidWorks: начертить тестовые 3D-модели. Сделать пару тестовых чертежей в КОМПАСе. Поддержание навыков.",
            "priority": "Средний",
            "deadline": "На этой неделе",
            "status": "Не начато",
            "fill": NORMAL_FILL_OP,
        },
        {
            "no": "6",
            "task": "Золотой набор данных: топ-20 деталей M3D+CDW",
            "details": "Подобрать 20 разнообразных деталей. Вручную проверить и утвердить пары M3D+CDW. Для тестирования этапов CAD-разработчика.",
            "priority": "Низкий",
            "deadline": "Апрель 2026",
            "status": "Не начато",
            "fill": LOW_FILL,
        },
    ]

    row = 5
    for t in tasks:
        ws.cell(row=row, column=1, value=t["no"])
        ws.cell(row=row, column=2, value=t["task"])
        ws.cell(row=row, column=3, value=t["details"])
        ws.cell(row=row, column=4, value=t["priority"])
        ws.cell(row=row, column=5, value=t["deadline"])
        ws.cell(row=row, column=6, value=t["status"])
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP
            if t.get("fill"):
                cell.fill = t["fill"]
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=6).alignment = CENTER
        ws.row_dimensions[row].height = 40
        row += 1

    # Ширина колонок
    widths = {"A": 5, "B": 45, "C": 60, "D": 12, "E": 18, "F": 14}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "C00000"

    return ws


# ================================================================
# MAIN
# ================================================================
def main():
    wb = openpyxl.Workbook()

    print("Создаю лист «План развития»...")
    create_plan_sheet(wb)

    print("Создаю лист «Журнал расходов»...")
    create_expenses_sheet(wb)

    print("Создаю лист «Бюджет»...")
    create_budget_sheet(wb)

    print("Создаю лист «Оперативные задачи»...")
    create_operative_sheet(wb)

    # Сохранение
    wb.save(OUTPUT_PATH)
    print(f"\nФайл сохранён: {OUTPUT_PATH}")
    print(f"Размер: {os.path.getsize(OUTPUT_PATH) / 1024:.1f} КБ")
    print(f"Листы: {wb.sheetnames}")


if __name__ == "__main__":
    main()
