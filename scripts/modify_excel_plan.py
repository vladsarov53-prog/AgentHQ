import shutil, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

src = r'D:\REDPEAK\Agent systems\AgentHQ\DATA\Дорожная карта проекта REDPEAK\RedPeak_Единый_план_2026_2028.xlsx'
dst = r'D:\REDPEAK\Agent systems\AgentHQ\DATA\Дорожная карта проекта REDPEAK\RedPeak_Единый_план_2026_2028_v2.xlsx'

shutil.copy2(src, dst)
wb = load_workbook(dst)

sheet_names = wb.sheetnames
ws_ops = wb[sheet_names[0]]
ws_plan = wb[sheet_names[1]]

# === Styles ===
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill('solid', fgColor='FFC00000')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
border_color = 'FFB4C6E7'
thin_border = Border(
    left=Side(style='thin', color=border_color),
    right=Side(style='thin', color=border_color),
    top=Side(style='thin', color=border_color),
    bottom=Side(style='thin', color=border_color)
)
data_font = Font(name='Calibri', size=10)
data_fill = PatternFill('solid', fgColor='FFFCE4D6')
data_align = Alignment(vertical='center', wrap_text=True)
data_align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# =============================================
# 1. ОПЕРАТИВНЫЕ ЗАДАЧИ - колонка G "Дата выполнения"
# =============================================

ws_ops.unmerge_cells('A1:F1')
ws_ops.unmerge_cells('A2:F2')
ws_ops.merge_cells('A1:G1')
ws_ops.merge_cells('A2:G2')

g4 = ws_ops.cell(row=4, column=7, value='Дата выполнения')
g4.font = header_font
g4.fill = header_fill
g4.alignment = header_align
g4.border = thin_border

ws_ops.column_dimensions['G'].width = 18.0
for row in range(5, ws_ops.max_row + 1):
    cell = ws_ops.cell(row=row, column=7)
    cell.font = data_font
    cell.fill = data_fill
    cell.alignment = data_align_center
    cell.border = thin_border
    cell.number_format = 'DD.MM.YYYY'

# Dropdown для статуса
dv_status = DataValidation(
    type='list',
    formula1='"В работе,Не начато,Выполнено,Завершено"',
    allow_blank=True
)
dv_status.error = 'Выберите статус из списка'
dv_status.errorTitle = 'Некорректный статус'
dv_status.prompt = 'Выберите статус задачи'
dv_status.promptTitle = 'Статус'
ws_ops.add_data_validation(dv_status)
dv_status.add('F5:F100')

# Условное форматирование: зеленый для "Выполнено"
green_fill_cf = PatternFill('solid', fgColor='FFC6EFCE')
ws_ops.conditional_formatting.add(
    'F5:F100',
    CellIsRule(operator='equal', formula=['"Выполнено"'], fill=green_fill_cf)
)

# =============================================
# 2. ПЛАН ПРОЕКТА - колонка I "Дата выполнения"
# =============================================

i4 = ws_plan.cell(row=4, column=9, value='Дата выполнения')
h4_cell = ws_plan.cell(row=4, column=8)
i4.font = copy.copy(h4_cell.font) if h4_cell.font else header_font
i4.fill = copy.copy(h4_cell.fill) if h4_cell.fill else header_fill
i4.alignment = copy.copy(h4_cell.alignment) if h4_cell.alignment else header_align
i4.border = copy.copy(h4_cell.border) if h4_cell.border else thin_border

ws_plan.column_dimensions['I'].width = 18.0
for row in range(5, ws_plan.max_row + 1):
    cell = ws_plan.cell(row=row, column=9)
    h_cell = ws_plan.cell(row=row, column=8)
    if h_cell.font:
        cell.font = copy.copy(h_cell.font)
    if h_cell.fill and h_cell.fill.patternType:
        cell.fill = copy.copy(h_cell.fill)
    cell.alignment = data_align_center
    if h_cell.border:
        cell.border = copy.copy(h_cell.border)
    cell.number_format = 'DD.MM.YYYY'

dv_plan = DataValidation(
    type='list',
    formula1='"В работе,Не начато,Выполнено,Завершено"',
    allow_blank=True
)
ws_plan.add_data_validation(dv_plan)
dv_plan.add('G5:G100')

ws_plan.conditional_formatting.add(
    'G5:G100',
    CellIsRule(operator='equal', formula=['"Выполнено"'], fill=green_fill_cf)
)

# =============================================
# 3. НОВЫЙ ЛИСТ "Выполненные задачи"
# =============================================

ws_done = wb.create_sheet('Выполненные задачи')

# Заголовок
ws_done.merge_cells('A1:H1')
title_cell = ws_done.cell(row=1, column=1, value='Выполненные задачи')
title_cell.font = Font(name='Calibri', bold=True, size=14, color='FF1F4E79')
title_cell.alignment = Alignment(horizontal='left', vertical='center')
ws_done.row_dimensions[1].height = 24

# Подзаголовок
ws_done.merge_cells('A2:H2')
ws_done.cell(row=2, column=1, value='Задачи со статусом "Выполнено" подтягиваются автоматически (FILTER). Обновляется при изменении статуса.')
ws_done.cell(row=2, column=1).font = Font(name='Calibri', size=9, italic=True, color='FF808080')
ws_done.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center')

# --- Секция 1: Оперативные задачи ---
ws_done.merge_cells('A4:H4')
sec1 = ws_done.cell(row=4, column=1, value='Оперативные задачи')
sec1.font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
sec1.fill = PatternFill('solid', fgColor='FFC00000')
sec1.alignment = Alignment(horizontal='left', vertical='center')
for c in range(2, 9):
    ws_done.cell(row=4, column=c).fill = PatternFill('solid', fgColor='FFC00000')
ws_done.row_dimensions[4].height = 22

# Заголовки строка 5
ops_headers = ['No', 'Задача', 'Детали / Результат', 'Приоритет', 'Дедлайн', 'Статус', 'Дата выполнения', 'Комментарий']
blue_fill = PatternFill('solid', fgColor='FF2E75B6')
for col, h in enumerate(ops_headers, 1):
    cell = ws_done.cell(row=5, column=col, value=h)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = header_align
    cell.border = thin_border

# FILTER формулы для оперативных задач
ops_sn = sheet_names[0]
for col_idx, src_col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G'], 1):
    cell = ws_done.cell(row=6, column=col_idx)
    cell.font = data_font
    cell.alignment = data_align if col_idx in [2, 3] else data_align_center
    cell.border = thin_border
    formula = f"=IFERROR(FILTER('{ops_sn}'!{src_col}5:{src_col}100,'{ops_sn}'!F5:F100=\"Выполнено\"),\"\")"
    cell.value = formula
    if col_idx == 7:
        cell.number_format = 'DD.MM.YYYY'

# Колонка H - пустая для комментариев
ws_done.cell(row=6, column=8).font = data_font
ws_done.cell(row=6, column=8).alignment = data_align
ws_done.cell(row=6, column=8).border = thin_border

# --- Секция 2: План проекта ---
ws_done.merge_cells('A22:I22')
sec2 = ws_done.cell(row=22, column=1, value='План проекта')
sec2.font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
sec2.fill = PatternFill('solid', fgColor='FFC00000')
sec2.alignment = Alignment(horizontal='left', vertical='center')
for c in range(2, 10):
    ws_done.cell(row=22, column=c).fill = PatternFill('solid', fgColor='FFC00000')
ws_done.row_dimensions[22].height = 22

# Заголовки строка 23
plan_headers = ['No', 'Веха', 'Задача', 'Исполнитель', 'Критерий', 'Дата', 'Статус', '% вып.', 'Дата выполнения']
for col, h in enumerate(plan_headers, 1):
    cell = ws_done.cell(row=23, column=col, value=h)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = header_align
    cell.border = thin_border

# FILTER формулы для плана проекта
plan_sn = sheet_names[1]
for col_idx, src_col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], 1):
    cell = ws_done.cell(row=24, column=col_idx)
    cell.font = data_font
    cell.alignment = data_align if col_idx == 3 else data_align_center
    cell.border = thin_border
    formula = f"=IFERROR(FILTER('{plan_sn}'!{src_col}5:{src_col}100,'{plan_sn}'!G5:G100=\"Выполнено\"),\"\")"
    cell.value = formula
    if col_idx == 9:
        cell.number_format = 'DD.MM.YYYY'

# Ширины колонок
ws_done.column_dimensions['A'].width = 6
ws_done.column_dimensions['B'].width = 35
ws_done.column_dimensions['C'].width = 50
ws_done.column_dimensions['D'].width = 16
ws_done.column_dimensions['E'].width = 22
ws_done.column_dimensions['F'].width = 16
ws_done.column_dimensions['G'].width = 14
ws_done.column_dimensions['H'].width = 16
ws_done.column_dimensions['I'].width = 18

# Закрепить строки
ws_done.freeze_panes = 'A6'

# --- Счётчики внизу ---
ws_done.cell(row=18, column=1, value='Итого оперативных:')
ws_done.cell(row=18, column=1).font = Font(name='Calibri', bold=True, size=10, color='FF1F4E79')
ws_done.cell(row=18, column=2).value = f"=COUNTIF('{ops_sn}'!F5:F100,\"Выполнено\")"
ws_done.cell(row=18, column=2).font = Font(name='Calibri', bold=True, size=10, color='FF1F4E79')

ws_done.cell(row=40, column=1, value='Итого по плану:')
ws_done.cell(row=40, column=1).font = Font(name='Calibri', bold=True, size=10, color='FF1F4E79')
ws_done.cell(row=40, column=2).value = f"=COUNTIF('{plan_sn}'!G5:G100,\"Выполнено\")"
ws_done.cell(row=40, column=2).font = Font(name='Calibri', bold=True, size=10, color='FF1F4E79')

wb.save(dst)
print(f'OK: {dst}')
